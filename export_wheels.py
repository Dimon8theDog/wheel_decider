#!/usr/bin/env python3
"""Export all 5 wheel configurations to a single .xlsx file."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Color palette (dark theme inspired)
HEADER_FILL = PatternFill("solid", fgColor="2D2D44")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DISABLED_FILL = PatternFill("solid", fgColor="3A3A4A")
DISABLED_FONT = Font(name="Calibri", color="888888", size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="7B68EE")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="AAAAAA")
PASS_FONT = Font(name="Calibri", bold=True, color="00CC66", size=11)
NUM_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="444444"),
    right=Side(style="thin", color="444444"),
    top=Side(style="thin", color="444444"),
    bottom=Side(style="thin", color="444444"),
)

wheels = [
    {
        "name": "Wheel 1 — €5",
        "target": 5,
        "undershoot_pct": 6.50,
        "ev": 4.67,
        "sectors": [
            ("15 FS",     1.50,  15.94, False),
            ("25 FS",     2.50,  15.55, False),
            ("50 FS",     5.00,  15.11, False),
            ("10 HB FS",  5.00,  14.59, False),
            ("€5",        5.00,  13.95, False),
            ("75 FS",     7.50,  13.10, False),
            ("15 HB FS",  7.50,  11.76, False),
            ("€10",      10.00,   0.00, True),
        ]
    },
    {
        "name": "Wheel 2 — €10",
        "target": 10,
        "undershoot_pct": 6.50,
        "ev": 9.35,
        "sectors": [
            ("75 FS",     7.50,  19.69, False),
            ("15 HB FS",  7.50,  18.78, False),
            ("100 FS",   10.00,  17.73, False),
            ("€10",      10.00,  16.47, False),
            ("20 HB FS", 10.00,  14.86, False),
            ("25 HB FS", 12.50,  12.47, False),
            ("€15",      15.00,   0.00, True),
            ("€20",      20.00,   0.00, True),
        ]
    },
    {
        "name": "Wheel 3 — €20",
        "target": 20,
        "undershoot_pct": 3.51,
        "ev": 19.30,
        "sectors": [
            ("25 HB FS", 12.50,  21.97, False),
            ("€15",      15.00,  20.23, False),
            ("175 FS",   17.50,  18.30, False),
            ("€20",      20.00,  16.09, False),
            ("€25",      25.00,  13.45, False),
            ("75 HB FS", 37.50,   9.96, False),
            ("€30",      30.00,   0.00, True),
            ("€35",      35.00,   0.00, True),
        ]
    },
    {
        "name": "Wheel 4 — €35",
        "target": 35,
        "undershoot_pct": 3.50,
        "ev": 33.78,
        "sectors": [
            ("€25",      25.00,  23.91, False),
            ("€30",      30.00,  21.36, False),
            ("€35",      35.00,  18.63, False),
            ("75 HB FS", 37.50,  15.64, False),
            ("€40",      40.00,  12.26, False),
            ("100 HB FS",50.00,   8.20, False),
            ("€60",      60.00,   0.00, True),
            ("€75",      75.00,   0.00, True),
        ]
    },
    {
        "name": "Wheel 5 — €65",
        "target": 65,
        "undershoot_pct": 4.00,
        "ev": 62.40,
        "sectors": [
            ("€25",      25.00,   0.00, True),
            ("€30",      30.00,   0.00, True),
            ("€35",      35.00,   0.00, True),
            ("€40",      40.00,   0.00, True),
            ("100 HB FS",50.00,  35.67, False),
            ("€60",      60.00,  29.05, False),
            ("€75",      75.00,  21.80, False),
            ("€80",      80.00,  13.48, False),
        ]
    },
]

def write_wheel_sheet(wb, wheel, is_first=False):
    if is_first:
        ws = wb.active
        ws.title = wheel["name"]
    else:
        ws = wb.create_sheet(title=wheel["name"])

    # Column widths
    col_widths = [6, 16, 12, 10, 12, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=wheel["name"])
    cell.font = TITLE_FONT
    row += 1

    # Subtitle with stats
    target = wheel["target"]
    ev = wheel["ev"]
    undershoot = wheel["undershoot_pct"]
    subtitle = f"Target: €{target}  |  EV: €{ev:.2f}  |  Undershoot: {undershoot:.2f}%"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=subtitle)
    cell.font = SUBTITLE_FONT
    row += 1

    # Rates reminder
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value="Rates: FS = €0.10/spin, HB FS = €0.50/spin")
    cell.font = Font(name="Calibri", size=10, color="999999", italic=True)
    row += 2

    # Headers
    headers = ["№", "Reward", "EUR", "Disabled", "Prob. %", "EV Contrib", "Cumul. EV"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    # Sector rows
    cumul_ev = 0
    for i, (label, eur, prob, disabled) in enumerate(wheel["sectors"], 1):
        ev_contrib = eur * prob / 100.0
        cumul_ev += ev_contrib

        row_font = DISABLED_FONT if disabled else NORMAL_FONT
        row_fill = DISABLED_FILL if disabled else PatternFill()

        # №
        cell = ws.cell(row=row, column=1, value=i)
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        # Reward
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = row_font
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # EUR
        cell = ws.cell(row=row, column=3, value=eur)
        cell.font = row_font
        cell.fill = row_fill
        cell.number_format = '€#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER

        # Disabled
        cell = ws.cell(row=row, column=4, value="FAKE" if disabled else "")
        cell.font = Font(name="Calibri", bold=True, color="FF6666", size=10) if disabled else row_font
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        # Prob %
        cell = ws.cell(row=row, column=5, value=prob / 100.0 if prob > 0 else 0)
        cell.font = row_font
        cell.fill = row_fill
        cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER

        # EV Contrib
        cell = ws.cell(row=row, column=6, value=ev_contrib)
        cell.font = row_font
        cell.fill = row_fill
        cell.number_format = '€#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER

        # Cumul EV
        cell = ws.cell(row=row, column=7, value=cumul_ev)
        cell.font = row_font
        cell.fill = row_fill
        cell.number_format = '€#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER

        row += 1

    # Summary row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value=f"Total EV: €{ev:.2f}  |  Undershoot: {undershoot:.2f}% of €{target}")
    cell.font = PASS_FONT


# Also create a combined "Summary" sheet
def write_summary_sheet(wb, wheels):
    ws = wb.create_sheet(title="Summary", index=0)

    col_widths = [22, 10, 10, 12, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="Bonus Wheels — Summary")
    cell.font = TITLE_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="FS = €0.10/spin  |  HB FS = €0.50/spin  |  8 sectors per wheel")
    cell.font = SUBTITLE_FONT
    row += 2

    headers = ["Wheel", "Target", "EV", "Undershoot", "Active", "Fake"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    for w in wheels:
        n_active = sum(1 for _, _, _, d in w["sectors"] if not d)
        n_fake = 8 - n_active

        ws.cell(row=row, column=1, value=w["name"]).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER

        c = ws.cell(row=row, column=2, value=w["target"])
        c.number_format = '€#,##0'; c.font = NORMAL_FONT; c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="right")

        c = ws.cell(row=row, column=3, value=w["ev"])
        c.number_format = '€#,##0.00'; c.font = NORMAL_FONT; c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="right")

        c = ws.cell(row=row, column=4, value=w["undershoot_pct"] / 100.0)
        c.number_format = '0.00%'; c.font = NORMAL_FONT; c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="right")

        c = ws.cell(row=row, column=5, value=n_active)
        c.font = NORMAL_FONT; c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=row, column=6, value=n_fake)
        c.font = NORMAL_FONT; c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

        row += 1


# Build workbook
for i, w in enumerate(wheels):
    write_wheel_sheet(wb, w, is_first=(i == 0))

write_summary_sheet(wb, wheels)

# Move summary to front
wb.move_sheet("Summary", offset=-len(wheels))

out = "/home/user/wheel_decider/bonus_wheels.xlsx"
wb.save(out)
print(f"Saved to {out}")
